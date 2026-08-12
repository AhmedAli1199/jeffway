"""
Navigate to EasyBOP Scheduling Register report and click Export (#btn_export_list).

https://easybop.co.uk/a_planned_works/z_works_reports/scheduling_register.php?contract_id=...
"""

from __future__ import annotations

import base64
import logging
import subprocess
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Awaitable, Callable, Optional, Tuple

from playwright.async_api import BrowserContext, Download, Page

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore

log = logging.getLogger("easybop.scheduling_register_export")

EXPORT_BUTTON_ID = "btn_export_list"
EXPORT_BUTTON_SELECTOR = f"#{EXPORT_BUTTON_ID}"
EXPORT_BUTTON_TEXT = "Export"
CANONICAL_EXPORT_FILENAME = "Scheduling-Register-Export.xlsx"
DEBUG_SCREENSHOT_PREFIX = "debug-scheduling-register-export"

_BUTTON_READY_JS = """() => {
  const btn = document.getElementById('btn_export_list');
  if (!btn) return false;
  const style = window.getComputedStyle(btn);
  if (style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') {
    return false;
  }
  const rect = btn.getBoundingClientRect();
  if (rect.width <= 0 || rect.height <= 0) return false;
  return !btn.disabled && !btn.classList.contains('ui-state-disabled');
}"""

_PAGE_STATE_JS = """() => {
  const lb = document.querySelector('#login_box');
  if (lb && lb.offsetParent !== null) return 'login';
  if (document.getElementById('btn_export_list')) return 'scheduling_register';
  const h1 = document.querySelector('h1');
  const h1txt = ((h1 && h1.textContent) || '').toLowerCase();
  if (h1txt.includes('easybop') && h1txt.includes('login')) return 'login';
  const body = (document.body && document.body.innerText) || '';
  if (body.includes('EasyBOP Login')) return 'login';
  return null;
}"""


def default_output_dir() -> Path:
    return Path(__file__).resolve().parent


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value.hour or value.minute or value.second else value.strftime("%Y-%m-%d")
    return str(value).strip()


def _cell_to_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value.hour or value.minute or value.second else value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return None
    return text


def _find_header_row(rows: list[tuple]) -> int:
    """EasyBOP report exports often have title rows before real headers."""
    hints = (
        ("client reference", "appointment date"),
        ("our reference", "appointment date"),
        ("client order reference", "appointment"),
        ("property reference", "appointment"),
        ("address", "appointment date"),
        ("board name", "operative name"),
    )
    for idx, row in enumerate(rows):
        cells = [_cell_to_str(c) for c in row]
        joined = " | ".join(cells).lower()
        for a, b in hints:
            if a in joined and b in joined:
                return idx
        if cells and cells[0].lower() in {
            "client order reference",
            "property reference",
            "address",
            "appointment date",
        }:
            return idx
    return 0


def _is_ole_xls(path: Path) -> bool:
    try:
        with path.open("rb") as f:
            sig = f.read(4)
        return sig == b"\xd0\xcf\x11\xe0"
    except Exception:
        return False


def _xls_to_xlsx(path: Path) -> Path:
    """EasyBOP scheduling register export is legacy OLE .xls — convert via LibreOffice."""
    lo = shutil.which("libreoffice") or shutil.which("soffice")
    if not lo:
        raise RuntimeError("LibreOffice required to convert scheduling register .xls export")

    out_dir = path.parent
    xlsx_path = path.with_suffix(".xlsx")
    if path.suffix.lower() == ".xlsx" and _is_ole_xls(path):
        raw = path.with_suffix(".xls")
        path.rename(raw)
        path = raw

    proc = subprocess.run(
        [lo, "--headless", "--convert-to", "xlsx", "--outdir", str(out_dir), str(path)],
        capture_output=True,
        text=True,
        timeout=120,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            f"LibreOffice xls→xlsx failed (exit {proc.returncode}): {proc.stderr or proc.stdout}"
        )

    converted = out_dir / f"{path.stem}.xlsx"
    if not converted.exists():
        raise RuntimeError(f"LibreOffice did not produce {converted}")

    if converted != xlsx_path:
        if xlsx_path.exists():
            xlsx_path.unlink()
        converted.replace(xlsx_path)

    if path.suffix.lower() == ".xls" and path.exists():
        path.unlink(missing_ok=True)

    log.info("scheduling register export: converted .xls → %s", xlsx_path.name)
    return xlsx_path


def normalize_scheduling_register_export_xlsx(path: Path) -> Path:
    """Rewrite export so row 1 is the true header row (for n8n headerRow: true)."""
    if _is_ole_xls(path):
        path = _xls_to_xlsx(path)

    if Workbook is None:
        log.warning("scheduling register export: openpyxl missing — skipping normalize")
        return path

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    if not all_rows:
        return path

    header_idx = _find_header_row(all_rows)
    header = [_cell_to_str(c) for c in all_rows[header_idx]]
    while header and not header[-1]:
        header.pop()
    if not header:
        return path

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Scheduling Register"
    out_ws.append(header)

    width = len(header)
    for row in all_rows[header_idx + 1 :]:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        values = [_cell_to_value(row[i]) if i < len(row) else None for i in range(width)]
        if not any(v is not None for v in values):
            continue
        out_ws.append(values)

    tmp = path.with_suffix(".normalized.xlsx")
    out_wb.save(tmp)
    tmp.replace(path)
    log.info(
        "scheduling register export: normalized — header row %s, %s data rows, %s cols",
        header_idx + 1,
        out_ws.max_row - 1,
        width,
    )
    return path


async def _on_login_screen(page: Page) -> bool:
    try:
        lb = page.locator("#login_box")
        if await lb.count() > 0 and await lb.first.is_visible():
            return True
    except Exception:
        pass
    if "easybop login" in (await page.title() or "").lower():
        return True
    return await page.get_by_text("EasyBOP Login", exact=False).count() > 0


async def _save_debug_screenshot(page: Page, out_dir: Path, label: str) -> Optional[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = out_dir / f"{DEBUG_SCREENSHOT_PREFIX}-{label}-{ts}.png"
    try:
        cdp = await page.context.new_cdp_session(page)
        shot = await cdp.send("Page.captureScreenshot", {"format": "png", "fromSurface": True})
        path.write_bytes(base64.b64decode(shot["data"]))
    except Exception as cdp_exc:
        log.warning("scheduling register: CDP screenshot failed (%s)", cdp_exc)
        try:
            await page.screenshot(path=str(path), full_page=True, timeout=8000, animations="disabled")
        except Exception:
            return None
    log.error("scheduling register: debug screenshot %s url=%s", path, page.url)
    return path


async def _navigate(page: Page, url: str, timeout_ms: int) -> None:
    log.info("scheduling register: navigating to %s", url)
    try:
        await page.goto(url, wait_until="commit", timeout=timeout_ms)
    except Exception as exc:
        log.warning("scheduling register: commit nav issue (%s), retry load", exc)
        try:
            await page.goto(url, wait_until="load", timeout=timeout_ms)
        except Exception as exc2:
            log.warning("scheduling register: load nav issue (%s)", exc2)


async def _wait_for_page(page: Page, timeout_ms: int) -> str:
    handle = await page.wait_for_function(_PAGE_STATE_JS, timeout=timeout_ms)
    state = await handle.json_value()
    log.info("scheduling register: state=%r url=%s", state, page.url)
    return state


async def _ensure_authenticated(
    page: Page,
    *,
    context: BrowserContext,
    url: str,
    login: Callable[[Page, str, str], Awaitable[bool]],
    username: str,
    password: str,
    session_path: Path,
    timeout_ms: int,
) -> None:
    for attempt in range(2):
        await _navigate(page, url, timeout_ms)
        state = await _wait_for_page(page, timeout_ms)
        if state == "scheduling_register":
            return
        if state == "login":
            log.info("scheduling register: login (attempt %s)", attempt + 1)
            ok = await login(page, username, password)
            if not ok:
                raise RuntimeError("EasyBOP login failed")
            await context.storage_state(path=str(session_path))
            continue
        raise RuntimeError(f"Unexpected page state: {state!r}")
    if await _on_login_screen(page):
        raise RuntimeError("Still on login after sign-in")


async def _wait_for_export_button(page: Page, timeout_ms: int):
    await page.wait_for_selector(EXPORT_BUTTON_SELECTOR, state="attached", timeout=timeout_ms)
    await page.wait_for_function(_BUTTON_READY_JS, timeout=timeout_ms)
    btn = page.locator(EXPORT_BUTTON_SELECTOR).first
    if await btn.count() == 0:
        btn = page.locator(f"button:has-text('{EXPORT_BUTTON_TEXT}')").first
    if await btn.count() == 0:
        raise RuntimeError("Export button not found on scheduling register")
    await btn.wait_for(state="visible", timeout=timeout_ms)
    await btn.scroll_into_view_if_needed(timeout=timeout_ms)
    return btn


async def _click_export(page: Page, btn, timeout_ms: int) -> Download:
    async with page.expect_download(timeout=timeout_ms) as download_info:
        try:
            await btn.click(timeout=timeout_ms)
        except Exception as exc:
            clicked = await page.evaluate(
                """() => {
                  const btn = document.getElementById('btn_export_list');
                  if (!btn) return false;
                  btn.click();
                  return true;
                }"""
            )
            if not clicked:
                raise RuntimeError("Could not click export button") from exc
    return await download_info.value


async def download_scheduling_register_export(
    page: Page,
    *,
    context: BrowserContext,
    base_url: str,
    contract_id: str,
    login: Callable[[Page, str, str], Awaitable[bool]],
    username: str,
    password: str,
    session_path: Path,
    timeout_ms: int,
    output_dir: Optional[Path] = None,
    report_year: Optional[int] = None,
    report_month: Optional[int] = None,
) -> Tuple[Path, str]:
    out = output_dir or default_output_dir()
    out.mkdir(parents=True, exist_ok=True)

    url = (
        f"{base_url.rstrip('/')}/a_planned_works/z_works_reports/"
        f"scheduling_register.php?contract_id={contract_id}"
    )
    try:
        await _ensure_authenticated(
            page,
            context=context,
            url=url,
            login=login,
            username=username,
            password=password,
            session_path=session_path,
            timeout_ms=timeout_ms,
        )
        if report_month is not None:
            year = report_year or datetime.now().year
            try:
                stage5_dir = Path(__file__).resolve().parent.parent / "Stage5"
                if str(stage5_dir) not in sys.path:
                    import sys as _sys

                    _sys.path.insert(0, str(stage5_dir))
                from month_filter import select_easybop_month  # type: ignore

                selected = await select_easybop_month(
                    page, year, report_month, timeout_ms=timeout_ms
                )
                if selected:
                    log.info(
                        "scheduling register: month filter set to %s-%02d",
                        year,
                        report_month,
                    )
                else:
                    log.warning(
                        "scheduling register: month filter not applied (%s-%02d)",
                        year,
                        report_month,
                    )
            except Exception as exc:
                log.warning("scheduling register: month filter error: %s", exc)
        btn = await _wait_for_export_button(page, timeout_ms)
        download = await _click_export(page, btn, timeout_ms)
        suggested = download.suggested_filename or "Scheduling-Register-Export.xls"
        # EasyBOP always suggests the same filename, so per-month downloads must be given
        # unique names or a later month overwrites the earlier one (losing that month's data).
        if report_month is not None:
            year = report_year or datetime.now().year
            stem = Path(suggested).stem
            ext = Path(suggested).suffix or ".xls"
            suggested = f"{stem}-{year}-{report_month:02d}{ext}"
        raw_dest = out / suggested
        await download.save_as(str(raw_dest))
        dest = raw_dest
        if _is_ole_xls(dest):
            dest = _xls_to_xlsx(dest)
        elif dest.suffix.lower() != ".xlsx":
            dest = dest.with_suffix(".xlsx")
            if raw_dest.exists() and raw_dest != dest:
                raw_dest.rename(dest)
        normalize_scheduling_register_export_xlsx(dest)
        log.info("scheduling register export: saved %s (%s bytes)", dest, dest.stat().st_size)
        return dest.resolve(), page.url
    except Exception as exc:
        label = type(exc).__name__.lower().replace("error", "") or "error"
        await _save_debug_screenshot(page, out, label)
        raise
