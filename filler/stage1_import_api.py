"""
Stage 1 import API — download WI/BOQ xlsx from Google Sheets and upload to EasyBOP.

Used locally (port 8001) + ngrok from n8n; VPS keeps template download endpoints.
"""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel, Field

_here = Path(__file__).resolve().parent
_repo = _here.parent
import sys

if str(_here) not in sys.path:
    sys.path.insert(0, str(_here))

logger = logging.getLogger("easybop.stage1_import")

# Stage 1 spreadsheet — Work Inst Import tab (legacy workbook kept for BOQ)
STAGE1_WORKBOOK_ID = "1ddliNGpjMaMsHh5V_0nROqBDsgJx7-_tq6wMMAC5qrg"
STAGE1_WI_IMPORT_SHEET_NAME = "Work Inst Import"
# Optional env override for anonymous Google export (?format=xlsx&gid=…)
STAGE1_WI_IMPORT_GID: Optional[str] = os.environ.get("STAGE1_WI_IMPORT_GID") or None

# Fixed Work Inst Import tab columns (order matters for EasyBOP import template).
WI_IMPORT_COLUMNS = [
    "Type of Work Code",
    "Description of Works",
    "Client Order Reference",
    "Issued By",
    "Recorded By",
    "Date Time Received (DD/MM/YYYY HH:MM)",
    "UPRN",
    "UPRN Alternative",
    "Resident Type (Leaseholder,Tenant)",
    "Occupancy (Occupied,Void)",
    "No. of Bedrooms",
    "Property Type",
    "Zone",
    "Client Estate Reference",
    "Client Property Status",
    "Year of Build",
    "Resident Title",
    "Resident First Name",
    "Resident Last Name",
    "Resident Tel",
    "Resident Mobile",
    "Resident Email",
    "Address 1",
    "Address 2",
    "Address 3",
    "Address 4",
    "town",
    "County",
    "Postcode",
    "Country",
    "Alternative Address 1",
    "Alternative Address 2",
    "Alternative Address 3",
    "Alternative Address 4",
    "Alternative Town",
    "Alternative County",
    "Alternative Postcode",
    "Alternative Country",
    "Vulnerability",
    "Vulnerability Code",
    "Vulnerability Description",
    "Handover date",
    "Next appointment date",
    "BCC Tracker Status",
    "Surveyor",
    "Clean Booked",
    "App number",
]

WI_WORKBOOK_ID = STAGE1_WORKBOOK_ID
WI_SHEET_GID: Optional[str] = STAGE1_WI_IMPORT_GID
LEGACY_WI_WORKBOOK_ID = "17WI7AM3bdtdwbMnBj4Ldcb7PVBv9iMOhoNxnqvdau3A"
LEGACY_WI_SHEET_GID = "1897422956"  # Reactive Works to Import- copy
BOQ_UPLOAD_COLUMNS = [
    "Works Instruction Ref.",
    "Works Instruction Address",
    "BOQ Item Name",
    "BOQ Item Section Name",
    "BOQ Item Ref.",
    "BOQ Item Short Description",
    "BOQ Item Unit",
    "BOQ Item Rate",
    "Quantity",
    "Location",
    "Notes",
]

STAGE1_BOQ_IMPORT_SHEET = "BOQ Import"
BOQ_SHEET_GID = "2075745851"  # legacy BOQ tab gid (fallback)
WI_DATETIME_COL = "Date Time Received (DD/MM/YYYY HH:MM)"
WI_DATETIME_FMT = "%d/%m/%Y %H:%M"
WI_DATETIME_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})(?::\d{2})?$")


def excel_downloads_dir() -> Path:
    d = _repo / "excel_downloads"
    d.mkdir(parents=True, exist_ok=True)
    return d


def resolve_excel_path(file_path: str) -> Path:
    """Map n8n/VPS paths like /root/.../WI-fill.xlsx to local excel_downloads/."""
    raw = (file_path or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="file_path is required")
    name = Path(raw.replace("\\", "/")).name
    local = excel_downloads_dir() / name
    p = Path(raw)
    if p.is_file():
        return p.resolve()
    if local.is_file():
        return local.resolve()
    return local.resolve()


async def download_sheet_xlsx(
    workbook_id: str,
    dest: Path,
    *,
    gid: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> tuple[int, bool]:
    from urllib.parse import quote

    if sheet_name:
        url = (
            f"https://docs.google.com/spreadsheets/d/{workbook_id}/export"
            f"?format=xlsx&sheet={quote(sheet_name)}"
        )
    elif gid:
        url = f"https://docs.google.com/spreadsheets/d/{workbook_id}/export?format=xlsx&gid={gid}"
    else:
        raise HTTPException(status_code=400, detail="Provide gid or sheet_name for export")
    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        resp = await client.get(url)
        if resp.status_code != 200:
            raise HTTPException(
                status_code=502,
                detail=f"Google Sheets export failed ({resp.status_code}) for gid={gid}. "
                "Ensure the sheet is shared as 'Anyone with the link can view'.",
            )
        content = resp.content
        if len(content) < 500 or content[:2] != b"PK":
            raise HTTPException(
                status_code=502,
                detail="Export did not return a valid xlsx (check sheet sharing / gid).",
            )
        dest.parent.mkdir(parents=True, exist_ok=True)
        existed = dest.exists()
        dest.write_bytes(content)
        return len(content), existed


def wi_row_has_corf(row: Dict[str, Any]) -> bool:
    corf = str(row.get("Client Order Reference") or "").strip()
    if not corf or re.match(r"^client order ref", corf, re.I):
        return False
    return True


def build_wi_fill_xlsx_from_rows(rows: List[Dict[str, Any]], dest: Path) -> int:
    """Write WI-fill.xlsx from sheet row dicts (header row + data in fixed column order)."""
    from openpyxl import Workbook

    pending = [r for r in rows if wi_row_has_corf(r)]
    if not pending:
        raise HTTPException(
            status_code=400,
            detail="No pending work instruction rows (Client Order Reference missing).",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Inst Import"

    for col_idx, col_name in enumerate(WI_IMPORT_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(pending, start=2):
        for col_idx, col_name in enumerate(WI_IMPORT_COLUMNS, start=1):
            val = row.get(col_name)
            if val is None:
                continue
            s = str(val).strip()
            if s == "":
                continue
            ws.cell(row=row_idx, column=col_idx, value=val)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    wb.close()
    return len(pending)


def boq_row_is_valid(row: Dict[str, Any]) -> bool:
    ref = str(row.get("BOQ Item Ref.") or "").strip()
    wi = str(row.get("Works Instruction Ref.") or "").strip()
    if not ref or not wi:
        return False
    m = re.match(r"^(\d+)/(\d+)$", wi)
    if not m:
        return False
    return int(m.group(1)) < 10000


def build_boq_xlsx_from_rows(rows: List[Dict[str, Any]], dest: Path) -> int:
    from openpyxl import Workbook

    pending = [r for r in rows if boq_row_is_valid(r)]
    if not pending:
        raise HTTPException(
            status_code=400,
            detail="No valid BOQ rows (need Works Instruction Ref + BOQ Item Ref).",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = STAGE1_BOQ_IMPORT_SHEET

    for col_idx, col_name in enumerate(BOQ_UPLOAD_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(pending, start=2):
        for col_idx, col_name in enumerate(BOQ_UPLOAD_COLUMNS, start=1):
            val = row.get(col_name)
            if val is None:
                continue
            s = str(val).strip()
            if s == "":
                continue
            ws.cell(row=row_idx, column=col_idx, value=val)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    wb.close()
    return len(pending)


def normalize_wi_datetime_value(value: Any) -> Any:
    """Normalize to strict DD/MM/YYYY HH:MM (no seconds, single space)."""
    if value is None or value == "":
        return value
    if isinstance(value, datetime):
        return value.strftime(WI_DATETIME_FMT)
    s = re.sub(r"\s+", " ", str(value).strip())
    m = WI_DATETIME_RE.match(s)
    if not m:
        return s
    day, month, year, hour, minute = (int(x) for x in m.groups())
    return datetime(year, month, day, hour, minute).strftime(WI_DATETIME_FMT)


def format_wi_fill_datetime_column(dest: Path) -> bool:
    """Normalize and right-align Date Time Received column in WI-fill.xlsx."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    wb = load_workbook(dest)
    ws = wb.active
    right = Alignment(horizontal="right")

    col_idx = None
    for cell in ws[1]:
        if cell.value and str(cell.value).strip() == WI_DATETIME_COL:
            col_idx = cell.column
            break

    if col_idx is None:
        logger.warning("WI-fill: column %r not found in header row", WI_DATETIME_COL)
        wb.close()
        return False

    normalized = 0
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.alignment = right
        if row == 1:
            continue
        if cell.value is None or cell.value == "":
            continue
        new_val = normalize_wi_datetime_value(cell.value)
        if new_val != cell.value:
            cell.value = new_val
            normalized += 1

    wb.save(dest)
    wb.close()
    logger.info(
        "WI-fill: formatted column %r (col %s, normalized %s rows)",
        WI_DATETIME_COL,
        col_idx,
        normalized,
    )
    return True


class DownloadExcelFilesRequest(BaseModel):
    download_wi: bool = Field(False, description="Save WI-fill.xlsx from pending WI sheet")
    download_boq: bool = Field(False, description="Save BOQ-verify.xlsx from BOQ tab")


class SaveWiFillRowsRequest(BaseModel):
    rows: List[Dict[str, Any]] = Field(
        ...,
        description="Work Inst Import row objects (from n8n Google Sheets read)",
    )


class SaveBoqRowsRequest(BaseModel):
    rows: List[Dict[str, Any]] = Field(
        ...,
        description="BOQ Import row objects ready for EasyBOP upload",
    )


class CreateWorksInstructionsRequest(BaseModel):
    contract_id: str = Field("321129", description="EasyBOP contract id")
    rows: List[Dict[str, Any]] = Field(
        ...,
        description="Work Inst Import rows — one Works Instruction created per CORF",
    )
    max_concurrent: int = Field(
        3,
        ge=1,
        le=3,
        description="Max parallel Playwright browsers (1–3)",
    )


class UploadFileRequest(BaseModel):
    contract_id: str = Field("321129", description="EasyBOP contract id (for logging)")
    file_path: str = Field(
        ...,
        description="Path to xlsx — VPS path ok; filename is resolved under excel_downloads/",
    )


class InsertBoqItemRequest(BaseModel):
    works_instruction_address: str = Field(
        ...,
        description="Works Instruction Address to search in EasyBOP (e.g. '29 HONEY GARSTON ROAD, BS13 9LY')",
    )
    boq_item_ref: str = Field(
        ...,
        description="BOQ Item Ref. to search for inside the JW Rates panel (e.g. '240251')",
    )
    quantity: float = Field(
        ...,
        description="Quantity to enter for the BOQ item",
        gt=0,
    )
    contract_id: str = Field("321129", description="EasyBOP contract_id")


def create_router(*, headless: bool) -> APIRouter:
    from insert_boq_item import insert_boq_item
    from upload_boq_verify import upload_boq_verify
    from upload_wi_fill import upload_wi_fill

    router = APIRouter(tags=["Stage 1 Import"])

    @router.post(
        "/download-excel-files",
        summary="Download WI-fill.xlsx and/or BOQ-verify.xlsx (overwrites excel_downloads/)",
    )
    async def download_excel_files(req: DownloadExcelFilesRequest) -> Dict[str, Any]:
        if not req.download_wi and not req.download_boq:
            raise HTTPException(status_code=400, detail="Set download_wi and/or download_boq true")

        out_dir = excel_downloads_dir()
        files: Dict[str, Any] = {}

        if req.download_wi:
            dest = out_dir / "WI-fill.xlsx"
            if not WI_SHEET_GID:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "WI Google export requires STAGE1_WI_IMPORT_GID (tab gid). "
                        "Use POST /save-wi-fill-rows with rows from n8n instead."
                    ),
                )
            nbytes, overwritten = await download_sheet_xlsx(
                WI_WORKBOOK_ID,
                dest,
                gid=WI_SHEET_GID,
            )
            datetime_aligned = format_wi_fill_datetime_column(dest)
            files["WI-fill.xlsx"] = {
                "path": str(dest),
                "bytes": nbytes,
                "overwritten": overwritten,
                "datetime_column_right_aligned": datetime_aligned,
            }
            logger.info(
                "Downloaded WI-fill.xlsx (%s bytes, overwrite=%s) -> %s",
                nbytes,
                overwritten,
                dest,
            )

        if req.download_boq:
            dest = out_dir / "BOQ-verify.xlsx"
            try:
                nbytes, overwritten = await download_sheet_xlsx(
                    STAGE1_WORKBOOK_ID,
                    dest,
                    sheet_name=STAGE1_BOQ_IMPORT_SHEET,
                )
            except HTTPException:
                nbytes, overwritten = await download_sheet_xlsx(
                    LEGACY_WI_WORKBOOK_ID,
                    dest,
                    gid=BOQ_SHEET_GID,
                )
            files["BOQ-verify.xlsx"] = {"path": str(dest), "bytes": nbytes, "overwritten": overwritten}
            logger.info(
                "Downloaded BOQ-verify.xlsx (%s bytes, overwrite=%s) -> %s",
                nbytes,
                overwritten,
                dest,
            )

        return {
            "success": True,
            "excel_downloads_dir": str(out_dir),
            "files": files,
        }

    @router.post(
        "/save-wi-fill-rows",
        summary="Build WI-fill.xlsx from Work Inst Import row data (preferred over Google export)",
    )
    async def save_wi_fill_rows(req: SaveWiFillRowsRequest) -> Dict[str, Any]:
        dest = excel_downloads_dir() / "WI-fill.xlsx"
        existed = dest.exists()
        row_count = build_wi_fill_xlsx_from_rows(req.rows, dest)
        datetime_aligned = format_wi_fill_datetime_column(dest)
        nbytes = dest.stat().st_size
        logger.info(
            "Built WI-fill.xlsx (%s bytes, %s rows, overwrite=%s) -> %s",
            nbytes,
            row_count,
            existed,
            dest,
        )
        return {
            "success": True,
            "path": str(dest),
            "bytes": nbytes,
            "row_count": row_count,
            "overwritten": existed,
            "datetime_column_right_aligned": datetime_aligned,
        }

    @router.post(
        "/save-boq-rows",
        summary="Build BOQ-verify.xlsx from BOQ Import row data",
    )
    async def save_boq_rows(req: SaveBoqRowsRequest) -> Dict[str, Any]:
        dest = excel_downloads_dir() / "BOQ-verify.xlsx"
        existed = dest.exists()
        row_count = build_boq_xlsx_from_rows(req.rows, dest)
        nbytes = dest.stat().st_size
        logger.info(
            "Built BOQ-verify.xlsx (%s bytes, %s rows, overwrite=%s) -> %s",
            nbytes,
            row_count,
            existed,
            dest,
        )
        return {
            "success": True,
            "path": str(dest),
            "bytes": nbytes,
            "row_count": row_count,
            "overwritten": existed,
        }

    @router.post(
        "/create-works-instructions",
        summary="Create Works Instructions via works_details.php form (one CORF per row)",
    )
    async def create_works_instructions_endpoint(req: CreateWorksInstructionsRequest) -> Dict[str, Any]:
        from create_works_instructions import create_works_instructions_batch

        if not req.rows:
            raise HTTPException(status_code=400, detail="rows is required")
        logger.info(
            "create-works-instructions contract_id=%s rows=%s max_concurrent=%s",
            req.contract_id,
            len(req.rows),
            req.max_concurrent,
        )
        try:
            return await create_works_instructions_batch(
                req.rows,
                contract_id=req.contract_id,
                headless=headless,
                max_concurrent=req.max_concurrent,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            logger.exception("create-works-instructions failed")
            raise HTTPException(status_code=500, detail=str(exc)) from exc

    @router.post("/upload-works-file", summary="Upload WI-fill.xlsx to EasyBOP import_works")
    async def upload_works_file(req: UploadFileRequest) -> Dict[str, Any]:
        path = resolve_excel_path(req.file_path)
        if not path.is_file():
            raise HTTPException(
                status_code=404,
                detail=f"File not found: {path}. Call POST /download-excel-files first.",
            )
        logger.info("upload-works-file contract_id=%s file=%s", req.contract_id, path)
        try:
            result = await upload_wi_fill(path, headless=headless, keep_open=False)
            return {"success": True, "contract_id": req.contract_id, **result}
        except Exception as exc:
            logger.exception("upload-works-file failed")
            raise HTTPException(status_code=500, detail=str(exc)) from exc

    @router.post("/upload-boq-file", summary="Upload BOQ-verify.xlsx to EasyBOP import_boqs")
    async def upload_boq_file(req: UploadFileRequest) -> Dict[str, Any]:
        path = resolve_excel_path(req.file_path)
        if not path.is_file():
            raise HTTPException(
                status_code=404,
                detail=f"File not found: {path}. Call POST /download-excel-files first.",
            )
        logger.info("upload-boq-file contract_id=%s file=%s", req.contract_id, path)
        try:
            result = await upload_boq_verify(path, headless=headless, keep_open=False)
            return {"success": True, "contract_id": req.contract_id, **result}
        except Exception as exc:
            logger.exception("upload-boq-file failed")
            raise HTTPException(status_code=500, detail=str(exc)) from exc

    @router.post(
        "/insert-boq-item",
        summary="Insert a single BOQ line into EasyBOP via the Works Instruction BOQ panel",
        description=(
            "Navigates to the Works Instruction index, searches by address, "
            "opens the BOQ tab → JW Rates panel, searches for the BOQ Item Ref, "
            "enters the quantity from the sheet, clicks Continue, then "
            "Save changes and Set Status → ACE Agreed → OK on the confirm dialog.\n\n"
            "On success the caller (n8n) should mark `IS_BOQ_INSERTED = true` "
            "in the **BOQ-unmatched-SOR** sheet."
        ),
    )
    async def insert_boq_item_endpoint(req: InsertBoqItemRequest) -> Dict[str, Any]:
        from playwright.async_api import async_playwright

        from config import settings as filler_settings
        from upload_wi_fill import resolve_session_path

        logger.info(
            "insert-boq-item: address=%r ref=%r qty=%s contract=%s",
            req.works_instruction_address,
            req.boq_item_ref,
            req.quantity,
            req.contract_id,
        )

        sp = resolve_session_path()

        try:
            async with async_playwright() as pw:
                browser = await pw.chromium.launch(
                    headless=headless,
                    slow_mo=filler_settings.slow_mo_ms or 0,
                    args=["--window-size=1400,900"],
                )
                ctx_kwargs: Dict[str, Any] = {}
                if sp.exists():
                    ctx_kwargs["storage_state"] = str(sp)
                context = await browser.new_context(**ctx_kwargs)
                page = await context.new_page()
                try:
                    result = await insert_boq_item(
                        page,
                        context,
                        works_instruction_address=req.works_instruction_address,
                        boq_item_ref=req.boq_item_ref,
                        quantity=req.quantity,
                        contract_id=req.contract_id,
                        timeout_ms=max(filler_settings.request_timeout_ms, 60_000),
                    )
                    await context.storage_state(path=str(sp))
                    return result
                finally:
                    await page.close()
                    await context.close()
                    await browser.close()
        except HTTPException:
            raise
        except Exception as exc:
            logger.exception("insert-boq-item failed")
            raise HTTPException(status_code=500, detail=str(exc)) from exc

    return router
