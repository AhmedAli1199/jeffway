"""Load appointment rows (Q7–Q11) from Google Sheet tab or local xlsx."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

_sheet_loader_mod = None


def _get_sheet_loader():
    global _sheet_loader_mod
    if _sheet_loader_mod is None:
        import importlib.util

        path = Path(__file__).resolve().parent / "sheet_loader.py"
        spec = importlib.util.spec_from_file_location("_appt_sheet_loader", path)
        if spec is None or spec.loader is None:
            raise ImportError(f"Could not load sheet_loader from {path}")
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _sheet_loader_mod = mod
    return _sheet_loader_mod

APPOINTMENT_SLOT_COLUMNS = [
    {"q_section": "7.1", "q_yn": "7.2", "trade": "Trade_1", "hours": "Hours_1", "notes": "Notes_1"},
    {"q_section": "8.1", "q_yn": "8.2", "trade": "Trade_2", "hours": "Hours_2", "notes": "Notes_2"},
    {"q_section": "9.1", "q_yn": "9.2", "trade": "Trade_3", "hours": "Hours_3", "notes": "Notes_3"},
    {"q_section": "10.1", "q_yn": "10.2", "trade": "Trade_4", "hours": "Hours_4", "notes": "Notes_4"},
    {"q_section": "11.1", "q_yn": "11.2", "trade": "Trade_5", "hours": "Hours_5", "notes": "Notes_5"},
]

DEFAULT_APPOINTMENT_GID = "1360938164"


def _cell_str(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    return s or None


def parse_uprn_from_q12(q12: Optional[str]) -> str:
    if not q12:
        return ""
    m = re.search(r"-\s*(\d{5,})\s*$", str(q12).strip())
    return m.group(1) if m else ""


def _row_dict_from_grid_row(header_map: Dict[str, int], data_row: List[Any]) -> Dict[str, Optional[str]]:
    out: Dict[str, Optional[str]] = {}
    for key, col_i in header_map.items():
        if key.startswith("__"):
            continue
        if col_i < len(data_row):
            v = _cell_str(data_row[col_i])
            if v is not None:
                out[key] = v
    return out


def load_appointment_row_from_google_url(
    google_sheet_url: str,
    *,
    uprn: Optional[str] = None,
    job_no: Optional[str] = None,
) -> Optional[Dict[str, Optional[str]]]:
    sl = _get_sheet_loader()
    sheet_id, gid = sl.parse_google_sheet_url(google_sheet_url)
    grid = sl.fetch_google_sheet_csv(sheet_id, gid)
    if len(grid) < 2:
        return None
    header_map = sl.build_header_map(grid[0])
    uprn_want = re.sub(r"\D", "", str(uprn or ""))
    job_want = (job_no or "").strip().lower()

    for data_row in grid[1:]:
        row = _row_dict_from_grid_row(header_map, list(data_row))
        row_uprn = re.sub(r"\D", "", str(row.get("UPRN") or ""))
        row_job = str(row.get("Job No") or row.get("Job No.") or "").strip().lower()
        if uprn_want and row_uprn == uprn_want:
            return row
        if job_want and row_job and row_job == job_want:
            return row
    return None


def load_appointment_row_from_xlsx(
    excel_path: Path,
    sheet_name: str,
    *,
    uprn: Optional[str] = None,
    job_no: Optional[str] = None,
) -> Optional[Dict[str, Optional[str]]]:
    path = excel_path.resolve()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return None
        header_map = _get_sheet_loader().build_header_map(rows[0])
        uprn_want = re.sub(r"\D", "", str(uprn or ""))
        job_want = (job_no or "").strip().lower()
        for data_row in rows[1:]:
            row = _row_dict_from_grid_row(header_map, list(data_row))
            row_uprn = re.sub(r"\D", "", str(row.get("UPRN") or ""))
            row_job = str(row.get("Job No") or row.get("Job No.") or "").strip().lower()
            if uprn_want and row_uprn == uprn_want:
                return row
            if job_want and row_job and row_job == job_want:
                return row
        return None
    finally:
        wb.close()


def appointment_slots_from_row(row: Dict[str, Optional[str]]) -> List[Dict[str, str]]:
    """Return list of non-empty appointment slots to fill, stopping after a 'no' on q_yn."""
    out: List[Dict[str, str]] = []
    for i, spec in enumerate(APPOINTMENT_SLOT_COLUMNS):
        if i > 0:
            prev_yn = (row.get(APPOINTMENT_SLOT_COLUMNS[i - 1]["q_yn"]) or "no").strip().lower()
            if prev_yn not in ("yes", "y", "true", "1"):
                break
        trade = (row.get(spec["trade"]) or "").strip()
        if not trade:
            break
        out.append(
            {
                "q_section": spec["q_section"],
                "q_yn": spec["q_yn"],
                "trade": trade,
                "hours": (row.get(spec["hours"]) or "").strip(),
                "notes": (row.get(spec["notes"]) or "").strip(),
                "another": (row.get(spec["q_yn"]) or "no").strip(),
            }
        )
    return out
