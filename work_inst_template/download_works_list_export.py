"""
Navigate to EasyBOP planned works index and click Export (#btn_export_list).

Uses the same BrowserContext session as the FastAPI app (cookies / storage_state).
"""

from __future__ import annotations

import base64
import logging
from datetime import datetime
from pathlib import Path
from typing import Awaitable, Callable, Optional, Tuple

from playwright.async_api import BrowserContext, Download, Page

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore

log = logging.getLogger("easybop.work_inst_list_export")

EXPORT_BUTTON_ID = "btn_export_list"
EXPORT_BUTTON_SELECTOR = f"#{EXPORT_BUTTON_ID}"
EXPORT_BUTTON_TEXT = "Export"
CANONICAL_EXPORT_FILENAME = "Works-List-Export.xlsx"
DEBUG_SCREENSHOT_PREFIX = "debug-works-list-export"

_BUTTON_READY_JS = """() => {
  const btn = document.getElementById('btn_export_list');
  if (!btn) return false;
  const style = window.getComputedStyle(btn);
  if (style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') {
    return false;
  }
  const rect = btn.getBoundingClientRect();
  if (rect.width <= 0 || rect.height <= 0) return false;
  return !btn.disabled && !btn.classList.contains('ui-state-disabled');
}"""

_PAGE_STATE_JS = """() => {
  const lb = document.querySelector('#login_box');
  if (lb && lb.offsetParent !== null) return 'login';
  if (document.getElementById('btn_export_list')) return 'works_index';
  const h1 = document.querySelector('h1');
  const h1txt = ((h1 && h1.textContent) || '').toLowerCase();
  if (h1txt.includes('easybop') && h1txt.includes('login')) return 'login';
  const body = (document.body && document.body.innerText) || '';
  if (body.includes('EasyBOP Login')) return 'login';
  return null;
}"""


def default_output_dir() -> Path:
    return Path(__file__).resolve().parent


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value.hour or value.minute or value.second else value.strftime("%Y-%m-%d")
    return str(value).strip()


def _cell_to_value(value):
    """Empty / blank cells → None (null), not empty string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value.hour or value.minute or value.second else value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return None
    return text


def _find_header_row(rows: list[tuple]) -> int:
    """EasyBOP export has title rows; real headers are around row 7."""
    for idx, row in enumerate(rows):
        cells = [_cell_to_str(c) for c in row]
        joined = " | ".join(cells).lower()
        if "received date and time" in joined and "appointment date" in joined:
            return idx
        if cells and cells[0].lower() == "received date and time":
            return idx
    return 0


def normalize_works_list_export_xlsx(path: Path) -> Path:
    """
    Rewrite export so row 1 is the true header row (for n8n headerRow: true).
    """
    if Workbook is None:
        log.warning("works list export: openpyxl missing — skipping xlsx normalize")
        return path

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    if not all_rows:
        return path

    header_idx = _find_header_row(all_rows)
    header = [_cell_to_str(c) for c in all_rows[header_idx]]
    # trim trailing empty headers
    while header and not header[-1]:
        header.pop()
    if not header:
        return path

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Works Instructions"
    out_ws.append(header)

    width = len(header)
    for row in all_rows[header_idx + 1 :]:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        values = [_cell_to_value(row[i]) if i < len(row) else None for i in range(width)]
        if not any(v is not None for v in values):
            continue
        out_ws.append(values)

    tmp = path.with_suffix(".normalized.xlsx")
    out_wb.save(tmp)
    tmp.replace(path)
    log.info(
        "works list export: normalized xlsx — header row was %s, %s data rows, %s columns",
        header_idx + 1,
        out_ws.max_row - 1,
        width,
    )
    return path


async def _on_login_screen(page: Page) -> bool:
    try:
        lb = page.locator("#login_box")
        if await lb.count() > 0 and await lb.first.is_visible():
            return True
    except Exception:
        pass
    if "easybop login" in (await page.title() or "").lower():
        return True
    return await page.get_by_text("EasyBOP Login", exact=False).count() > 0


async def _save_debug_screenshot(page: Page, out_dir: Path, label: str) -> Optional[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = out_dir / f"{DEBUG_SCREENSHOT_PREFIX}-{label}-{ts}.png"
    try:
        cdp = await page.context.new_cdp_session(page)
        shot = await cdp.send("Page.captureScreenshot", {"format": "png", "fromSurface": True})
        path.write_bytes(base64.b64decode(shot["data"]))
    except Exception as cdp_exc:
        log.warning("works list export: CDP screenshot failed (%s), trying page.screenshot", cdp_exc)
        try:
            await page.screenshot(path=str(path), full_page=True, timeout=8000, animations="disabled")
        except Exception as shot_exc:
            log.warning("works list export: could not save debug screenshot: %s", shot_exc)
            return None

    log.error(
        "works list export: debug screenshot saved %s (url=%s title=%r)",
        path,
        page.url,
        await page.title(),
    )
    return path


async def _navigate_to_works_index(page: Page, url: str, timeout_ms: int) -> None:
    log.info("works list export: navigating to %s", url)
    try:
        await page.goto(url, wait_until="commit", timeout=timeout_ms)
    except Exception as exc:
        log.warning("works list export: commit navigation issue (%s), retrying with load", exc)
        try:
            await page.goto(url, wait_until="load", timeout=timeout_ms)
        except Exception as exc2:
            log.warning("works list export: load navigation issue (%s)", exc2)
    log.info("works list export: navigation committed url=%s", page.url)


async def _wait_for_login_or_works_index(page: Page, timeout_ms: int) -> str:
    log.info("works list export: waiting for login screen or works index page")
    handle = await page.wait_for_function(_PAGE_STATE_JS, timeout=timeout_ms)
    state = await handle.json_value()
    log.info(
        "works list export: page state=%r url=%s title=%r",
        state,
        page.url,
        await page.title(),
    )
    return state


async def _ensure_authenticated_on_works_index(
    page: Page,
    *,
    context: BrowserContext,
    url: str,
    login: Callable[[Page, str, str], Awaitable[bool]],
    username: str,
    password: str,
    session_path: Path,
    timeout_ms: int,
) -> None:
    for attempt in range(2):
        await _navigate_to_works_index(page, url, timeout_ms)
        state = await _wait_for_login_or_works_index(page, timeout_ms)

        if state == "works_index":
            return

        if state == "login":
            log.info("works list export: session expired — signing in (attempt %s)", attempt + 1)
            ok = await login(page, username, password)
            if not ok:
                raise RuntimeError("EasyBOP login failed")
            await context.storage_state(path=str(session_path))
            continue

        raise RuntimeError(f"Unexpected page state after navigation: {state!r}")

    if await _on_login_screen(page):
        raise RuntimeError("Still on login after sign-in")


async def _wait_for_export_button(page: Page, timeout_ms: int):
    log.info("works list export: waiting for %s", EXPORT_BUTTON_SELECTOR)
    await page.wait_for_selector(EXPORT_BUTTON_SELECTOR, state="attached", timeout=timeout_ms)
    await page.wait_for_function(_BUTTON_READY_JS, timeout=timeout_ms)

    btn = page.locator(EXPORT_BUTTON_SELECTOR).first
    if await btn.count() == 0:
        btn = page.locator(f"button:has-text('{EXPORT_BUTTON_TEXT}')").first
    if await btn.count() == 0:
        raise RuntimeError(
            f"Export button not found: {EXPORT_BUTTON_SELECTOR!r} or text {EXPORT_BUTTON_TEXT!r}"
        )

    await btn.wait_for(state="visible", timeout=timeout_ms)
    await btn.scroll_into_view_if_needed(timeout=timeout_ms)
    log.info("works list export: export button ready")
    return btn


async def _click_export_button(page: Page, btn, timeout_ms: int) -> Download:
    async with page.expect_download(timeout=timeout_ms) as download_info:
        try:
            await btn.click(timeout=timeout_ms)
        except Exception as exc:
            log.warning("works list export: Playwright click failed (%s), using JS click", exc)
            clicked = await page.evaluate(
                """() => {
                  const btn = document.getElementById('btn_export_list');
                  if (!btn) return false;
                  btn.click();
                  return true;
                }"""
            )
            if not clicked:
                raise RuntimeError(f"Could not click export button: {EXPORT_BUTTON_SELECTOR}") from exc
    return await download_info.value


async def download_works_list_export(
    page: Page,
    *,
    context: BrowserContext,
    base_url: str,
    contract_id: str,
    login: Callable[[Page, str, str], Awaitable[bool]],
    username: str,
    password: str,
    session_path: Path,
    timeout_ms: int,
    output_dir: Optional[Path] = None,
) -> Tuple[Path, str]:
    """
    Open z_works/index.php, ensure session, click #btn_export_list, save download.

    Returns (absolute_path_saved, final_page_url).
    """
    out = output_dir or default_output_dir()
    out.mkdir(parents=True, exist_ok=True)

    url = f"{base_url.rstrip('/')}/a_planned_works/z_works/index.php?contract_id={contract_id}"
    try:
        await _ensure_authenticated_on_works_index(
            page,
            context=context,
            url=url,
            login=login,
            username=username,
            password=password,
            session_path=session_path,
            timeout_ms=timeout_ms,
        )

        btn = await _wait_for_export_button(page, timeout_ms)
        download = await _click_export_button(page, btn, timeout_ms)

        dest = out / CANONICAL_EXPORT_FILENAME
        await download.save_as(str(dest))
        normalize_works_list_export_xlsx(dest)
        size = dest.stat().st_size
        log.info("works list export: saved %s (%s bytes)", dest, size)
        return dest.resolve(), page.url
    except Exception as exc:
        label = type(exc).__name__.lower().replace("error", "")
        await _save_debug_screenshot(page, out, label or "error")
        raise
