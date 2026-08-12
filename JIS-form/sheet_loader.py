"""Load JIS / SOR columns from Sheet3 (xlsx) or a Google Sheets tab (CSV export)."""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook

JIS_COLUMN_KEYS: List[str] = [
    "Q1.1",
    "Q1.2",
    "Q1.3",
    "Q1.4",
    "Q1.5",
    "Q1.6",
    "Q2.1",
    "Q2.2",
    "Q3.1",
    "Q3.2",
    "Q4.1",
    "Q4.2",
    "Q5.1",
    "Q5.2",
    "QS-Task",
]

EXAMPLE_HEADER_ALIASES = frozenset(
    {
        "example #",
        "example#",
        "example no",
        "example no.",
        "example_number",
        "example",
    }
)

# Optional columns to disambiguate pw_works_show_pre_item_dialog when several rows share contract_id + wi_pw
PRE_ITEM_ITEM_ID_HEADERS = frozenset(
    {
        "item id",
        "item_id",
        "jis item id",
        "jis_item_id",
        "pre_item_item_id",
        "jis item",
    }
)
PRE_ITEM_WORKS_ID_HEADERS = frozenset(
    {
        "works id",
        "works_id",
        "jis_works_id",
        "pre_item_works_id",
    }
)

GOOGLE_EXPORT_CSV = "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def _norm_header(cell: Any) -> str:
    if cell is None:
        return ""
    s = " ".join(str(cell).split()).strip().lower()
    return s


def _strip_derived_annotation(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r",\s*\(derived\)\s*$", "", s, flags=re.IGNORECASE).strip()
    return s or None


def _cell_to_str(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if isinstance(value, bool):
        return "yes" if value else "no"
    s = str(value).strip()
    return _strip_derived_annotation(s)


def _example_cell_matches(cell: Any, want: int) -> bool:
    if cell is None:
        return False
    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
        try:
            return int(cell) == want
        except (ValueError, OverflowError):
            pass
    s = str(cell).strip()
    if not s:
        return False
    s = re.sub(r",\s*\(derived\)\s*$", "", s, flags=re.IGNORECASE).strip()
    try:
        return int(float(s)) == want
    except ValueError:
        return s == str(want)


def build_header_map(header_row: Sequence[Any]) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _norm_header(cell)
        if not key:
            continue
        if key in EXAMPLE_HEADER_ALIASES:
            header_map["__example__"] = idx
            continue
        if key in PRE_ITEM_ITEM_ID_HEADERS:
            header_map["__pre_item_item_id__"] = idx
            continue
        if key in PRE_ITEM_WORKS_ID_HEADERS:
            header_map["__pre_item_works_id__"] = idx
            continue
        qmatch = re.fullmatch(r"q\s*(\d+)\s*\.\s*(\d+)", key.replace(" ", ""))
        if qmatch:
            canon = f"Q{qmatch.group(1)}.{qmatch.group(2)}"
            header_map[canon] = idx
            continue
        for k in JIS_COLUMN_KEYS:
            if key.upper().replace(" ", "") == k.upper().replace(" ", ""):
                header_map[k] = idx
                break
    return header_map


def _row_to_jis_dict(header_map: Dict[str, int], data_row: Sequence[Any]) -> Dict[str, Optional[str]]:
    out: Dict[str, Optional[str]] = {}
    for k in JIS_COLUMN_KEYS:
        if k not in header_map:
            continue
        col_i = header_map[k]
        if col_i < len(data_row):
            v = _cell_to_str(data_row[col_i])
            if v is not None:
                out[k] = v
    return out


def extract_jis_row_for_example(
    header_row: Sequence[Any],
    data_rows: Iterable[Sequence[Any]],
    example_number: int,
    *,
    source_hint: str = "",
) -> Dict[str, Optional[str]]:
    header_map = build_header_map(header_row)
    if "__example__" not in header_map:
        sample = [header_row[i] for i in range(min(15, len(header_row)))]
        raise KeyError(
            f"No 'Example #' column found{f' in {source_hint}' if source_hint else ''}. "
            f"Header sample: {sample}"
        )

    ex_idx = header_map["__example__"]
    found: Optional[Sequence[Any]] = None
    for data_row in data_rows:
        row = list(data_row)
        if ex_idx < len(row) and _example_cell_matches(row[ex_idx], example_number):
            found = row
            break

    if found is None:
        raise LookupError(
            f"No row with Example # = {example_number}"
            f"{f' in {source_hint}' if source_hint else ''}"
        )

    out = _row_to_jis_dict(header_map, found)
    for meta_key, map_key in (
        ("pre_item_item_id", "__pre_item_item_id__"),
        ("pre_item_works_id", "__pre_item_works_id__"),
    ):
        if map_key not in header_map:
            continue
        ix = header_map[map_key]
        if ix < len(found):
            v = _cell_to_str(found[ix])
            if v:
                out[meta_key] = v
    return out


def parse_google_sheet_url(url: str) -> Tuple[str, str]:
    """
    Extract spreadsheet id and gid from a Google Sheets browser URL.
    """
    u = url.strip()
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", u)
    if not m:
        raise ValueError("Not a Google Sheets URL (expected /spreadsheets/d/<id>/...).")
    sheet_id = m.group(1)

    parsed = urlparse(u)
    qs = parse_qs(parsed.query)
    gid: Optional[str] = None
    if "gid" in qs and qs["gid"]:
        gid = qs["gid"][0]
    if gid is None and parsed.fragment:
        fm = re.search(r"gid=(\d+)", parsed.fragment)
        if fm:
            gid = fm.group(1)
    if not gid:
        raise ValueError("URL must include tab id ?gid=... (or #gid=...) in the link.")
    return sheet_id, gid


def fetch_google_sheet_csv_rows(sheet_id: str, gid: str, *, timeout_s: int = 60) -> List[List[str]]:
    url = GOOGLE_EXPORT_CSV.format(sheet_id=sheet_id, gid=gid)
    req = Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; JIS-form/1.0)"})
    try:
        with urlopen(req, timeout=timeout_s) as resp:
            raw = resp.read()
    except HTTPError as e:
        if e.code in (401, 403):
            raise ValueError(
                "Google Sheets blocked export (HTTP %s). Set the spreadsheet to "
                "'Anyone with the link can view', or download .xlsx and use excel_path."
                % e.code
            ) from e
        raise ValueError(f"Google Sheets export failed: HTTP {e.code}") from e
    except URLError as e:
        raise ValueError(f"Could not reach Google Sheets: {e}") from e

    text = raw.decode("utf-8-sig", errors="replace")
    if "<!doctype html" in text[:200].lower() or "<html" in text[:500].lower():
        raise ValueError(
            "Google returned HTML instead of CSV — document may be private. "
            "Use link sharing 'Anyone with the link can view', or use excel_path."
        )

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("Google Sheets CSV export is empty.")
    return rows


def load_jis_row_from_google_url(google_sheet_url: str, example_number: int) -> Dict[str, Optional[str]]:
    sheet_id, gid = parse_google_sheet_url(google_sheet_url)
    grid = fetch_google_sheet_csv_rows(sheet_id, gid)
    header = grid[0]
    return extract_jis_row_for_example(
        header,
        grid[1:],
        example_number,
        source_hint=f"Google Sheet {sheet_id} gid={gid}",
    )


def load_jis_rows_from_google_url(
    google_sheet_url: str,
    *,
    require_q13: bool = True,
) -> List[Dict[str, Optional[str]]]:
    """
    Load all JIS rows from a Google Sheet tab (e.g. gid=jis).
    Skips blank rows; when require_q13, only rows with Q1.3 populated.
    Each dict includes ``row_number`` (1-based sheet row, header = row 1).
    """
    sheet_id, gid = parse_google_sheet_url(google_sheet_url)
    grid = fetch_google_sheet_csv_rows(sheet_id, gid)
    if len(grid) < 2:
        return []
    header = grid[0]
    header_map = build_header_map(header)
    out: List[Dict[str, Optional[str]]] = []
    for row_idx, data_row in enumerate(grid[1:], start=2):
        row = list(data_row)
        vals = _row_to_jis_dict(header_map, row)
        if not vals:
            continue
        if require_q13 and not (vals.get("Q1.3") or "").strip():
            continue
        vals["row_number"] = row_idx
        out.append(vals)
    return out


def load_jis_row_for_example(
    excel_path: Path,
    sheet_name: str,
    example_number: int,
) -> Dict[str, Optional[str]]:
    """
    Return a dict of Q1.1..Q5.2 for the first row where the Example # column equals example_number.
    """
    path = excel_path.resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Sheet {sheet_name!r} not in workbook (have: {wb.sheetnames[:20]}"
                f"{'…' if len(wb.sheetnames) > 20 else ''})"
            )
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise ValueError("Sheet is empty")

        header_tuple = tuple(header_row) if header_row is not None else ()
        return extract_jis_row_for_example(
            header_tuple,
            rows,
            example_number,
            source_hint=f"xlsx {path.name} / {sheet_name!r}",
        )
    finally:
        wb.close()
