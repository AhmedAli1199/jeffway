#!/usr/bin/env python3
"""
Compare Scheduling Register export vs Works List export.

Match key: Client Order Reference (unique per address / variation).
Compute min future appointment date from scheduling register per CORF.
Diff against Works List 'Next appointment date' for active (not completed) jobs.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required", file=sys.stderr)
    sys.exit(1)

# Only this scheduling status should drive next-appointment updates.
BOOKED_STAFF_ALLOCATED_STATUS = "appointment booked - staff allocated"

WI_COL_ALIASES = {
    "client_order_reference": ("client order reference",),
    "property_reference": ("property reference",),
    "address": ("address",),
    "next_appointment_date": ("next appointment date",),
    "completed": ("completed",),
    "works_status": ("works status",),
    "description": ("description of work",),
}

SCHED_COL_ALIASES = {
    "client_order_reference": (
        "client order reference",
        "client reference",
        "corf",
        "client order ref",
    ),
    "property_reference": ("property reference", "property ref"),
    "address": ("address",),
    "appointment_date": (
        "appointment date",
        "appointment date/time",
        "appointment date and time",
        "date",
    ),
    "status": ("appointment status", "status", "scheduling status"),
}


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def _find_col(headers: List[str], aliases: Tuple[str, ...]) -> Optional[int]:
    normed = [_norm_header(h) for h in headers]
    for alias in aliases:
        alias_n = _norm_header(alias)
        for i, h in enumerate(normed):
            if h == alias_n or alias_n in h:
                return i
    return None


def _read_xlsx_rows(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    out: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        item = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            item[h] = val
        out.append(item)
    return headers, out


def _cell(row: Dict[str, Any], col_idx: Optional[int], headers: List[str]) -> str:
    if col_idx is None:
        return ""
    key = headers[col_idx]
    v = row.get(key)
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def _norm_corf(v: str) -> str:
    return re.sub(r"\s+", "", str(v or "").strip().upper())


def _is_completed(val: str) -> bool:
    s = str(val or "").strip().lower()
    return s in {"yes", "y", "true", "1", "completed"}


def _norm_status(status: str) -> str:
    s = _norm_header(status)
    s = re.sub(r"[–—−]", "-", s)
    return re.sub(r"\s*-\s*", " - ", s)


def _status_active(status: str) -> bool:
    return _norm_status(status) == BOOKED_STAFF_ALLOCATED_STATUS


def build_next_dates_from_scheduling(
    sched_rows: List[Dict[str, Any]],
    sched_headers: List[str],
    *,
    as_of: date,
    current_by_corf: Optional[Dict[str, Optional[date]]] = None,
) -> Dict[str, Dict[str, Any]]:
    """client_order_reference -> {next_date, address, property_ref, appointment_count}."""
    cols = {k: _find_col(sched_headers, v) for k, v in SCHED_COL_ALIASES.items()}
    by_corf: Dict[str, List[date]] = defaultdict(list)
    meta: Dict[str, Dict[str, Any]] = {}

    for row in sched_rows:
        corf = _norm_corf(_cell(row, cols["client_order_reference"], sched_headers))
        if not corf:
            continue
        status = _cell(row, cols["status"], sched_headers)
        if not _status_active(status):
            continue
        appt = _parse_date(_cell(row, cols["appointment_date"], sched_headers))
        if appt is None:
            appt = _parse_date(row.get(sched_headers[cols["appointment_date"]]) if cols["appointment_date"] is not None else None)
        if appt is None:
            continue
        by_corf[corf].append(appt)
        if corf not in meta:
            meta[corf] = {
                "client_order_reference": _cell(row, cols["client_order_reference"], sched_headers),
                "address": _cell(row, cols["address"], sched_headers),
                "property_reference": _cell(row, cols["property_reference"], sched_headers),
            }

    def _pick_next(corf: str, dates: List[date]) -> Optional[date]:
        if not dates:
            return None
        sorted_dates = sorted(dates)
        future = [d for d in sorted_dates if d >= as_of]
        if future:
            return future[0]
        current = (current_by_corf or {}).get(corf)
        if current:
            newer = [d for d in sorted_dates if d > current]
            if newer:
                return newer[0]
        return None

    result: Dict[str, Dict[str, Any]] = {}
    for corf, dates in by_corf.items():
        nxt = _pick_next(corf, dates)
        if nxt is None:
            continue
        result[corf] = {
            **meta.get(corf, {}),
            "computed_next_appointment_date": nxt.isoformat(),
            "future_appointment_count": len([d for d in dates if d >= as_of]),
        }
    return result


def compare_next_appointment(
    scheduling_path: Path,
    works_list_path: Path,
    *,
    as_of: Optional[date] = None,
) -> Dict[str, Any]:
    as_of = as_of or date.today()
    sched_headers, sched_rows = _read_xlsx_rows(scheduling_path)
    wi_headers, wi_rows = _read_xlsx_rows(works_list_path)

    wi_cols = {k: _find_col(wi_headers, v) for k, v in WI_COL_ALIASES.items()}

    current_by_corf: Dict[str, Optional[date]] = {}
    for row in wi_rows:
        corf = _norm_corf(_cell(row, wi_cols["client_order_reference"], wi_headers))
        if not corf or corf in current_by_corf:
            continue
        current_by_corf[corf] = _parse_date(_cell(row, wi_cols["next_appointment_date"], wi_headers))

    sched_next = build_next_dates_from_scheduling(
        sched_rows, sched_headers, as_of=as_of, current_by_corf=current_by_corf
    )

    updates: List[Dict[str, Any]] = []
    clear_blank: List[Dict[str, Any]] = []
    already_ok: List[Dict[str, Any]] = []
    no_sched_match: List[Dict[str, Any]] = []
    skipped_completed = 0

    for row in wi_rows:
        corf = _norm_corf(_cell(row, wi_cols["client_order_reference"], wi_headers))
        if not corf:
            continue
        completed = _cell(row, wi_cols["completed"], wi_headers)
        if _is_completed(completed):
            skipped_completed += 1
            continue

        current_raw = _cell(row, wi_cols["next_appointment_date"], wi_headers)
        current = _parse_date(current_raw)
        sched = sched_next.get(corf)

        base = {
            "client_order_reference": _cell(row, wi_cols["client_order_reference"], wi_headers),
            "property_reference": _cell(row, wi_cols["property_reference"], wi_headers),
            "address": _cell(row, wi_cols["address"], wi_headers),
            "description_of_work": _cell(row, wi_cols["description"], wi_headers),
            "works_status": _cell(row, wi_cols["works_status"], wi_headers),
            "current_next_appointment_date": current_raw or None,
        }

        if not sched:
            if current_raw:
                no_sched_match.append({**base, "action": "clear_or_investigate", "reason": "no_future_appointment_in_register"})
            else:
                no_sched_match.append({**base, "action": "already_blank", "reason": "no_future_appointment_in_register"})
            continue

        computed = sched["computed_next_appointment_date"]
        computed_d = _parse_date(computed)
        if current == computed_d:
            already_ok.append({**base, "computed_next_appointment_date": computed})
        elif computed_d:
            updates.append({
                **base,
                "computed_next_appointment_date": computed,
                "future_appointment_count": sched["future_appointment_count"],
                "action": "update",
            })

    return {
        "as_of": as_of.isoformat(),
        "scheduling_rows": len(sched_rows),
        "works_list_rows": len(wi_rows),
        "scheduling_headers": sched_headers,
        "works_list_headers": wi_headers,
        "scheduling_jobs_with_future_appt": len(sched_next),
        "skipped_completed_wi": skipped_completed,
        "needs_update": updates,
        "already_correct": already_ok,
        "no_future_in_register": no_sched_match,
        "summary": {
            "needs_update_count": len(updates),
            "already_correct_count": len(already_ok),
            "no_future_in_register_count": len(no_sched_match),
        },
    }


def main() -> None:
    p = argparse.ArgumentParser(description="Compare scheduling register vs works list for Next appointment date")
    p.add_argument("scheduling_xlsx")
    p.add_argument("works_list_xlsx")
    p.add_argument("--json", action="store_true", help="Print full JSON")
    p.add_argument("--limit", type=int, default=30, help="Max update rows to print")
    args = p.parse_args()

    result = compare_next_appointment(Path(args.scheduling_xlsx), Path(args.works_list_xlsx))
    if args.json:
        print(json.dumps(result, indent=2, default=str))
        return

    s = result["summary"]
    print(f"As of: {result['as_of']}")
    print(f"Scheduling rows: {result['scheduling_rows']} | WI rows: {result['works_list_rows']}")
    print(f"Scheduling headers: {result['scheduling_headers'][:12]}...")
    print(f"Jobs with future appt (by CORF): {result['scheduling_jobs_with_future_appt']}")
    print(f"Needs update: {s['needs_update_count']} | Already correct: {s['already_correct_count']} | No future in register: {s['no_future_in_register_count']}")
    print(f"Skipped completed WI: {result['skipped_completed_wi']}")
    print()
    if result["needs_update"]:
        print("--- NEEDS UPDATE (sample) ---")
        for row in result["needs_update"][: args.limit]:
            print(
                f"  {row.get('client_order_reference')} | {row.get('address', '')[:40]} | "
                f"current={row.get('current_next_appointment_date')!r} -> {row.get('computed_next_appointment_date')}"
            )
    if result["no_future_in_register"]:
        print()
        print("--- NO FUTURE APPOINTMENT IN REGISTER (sample, may need clear) ---")
        for row in result["no_future_in_register"][: min(10, args.limit)]:
            if row.get("current_next_appointment_date"):
                print(
                    f"  {row.get('client_order_reference')} | current={row.get('current_next_appointment_date')!r} | "
                    f"{row.get('address', '')[:40]}"
                )


if __name__ == "__main__":
    main()
